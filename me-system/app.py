from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import sqlite3
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

app = Flask(__name__)
CORS(app)

DB_PATH = os.path.join(os.path.dirname(__file__), 'me_system.db')

CATEGORIES = [
    '空調設備', '電力設備', '給排水設備', '消防設備',
    '弱電系統', '機械設備', '電梯/電扶梯', '其他'
]

IMPORT_COLUMNS = ['名稱*', '類別*', '年度*', '金額(元)*', '品牌', '型號', '規格', '單位', '數量', '專案', '供應商', '備註']

SAMPLE_DATA = [
    ('變頻離心式冰水主機', '空調設備', '約克', 'YLAA0470', '1000RT 380V 60Hz', '台', 2023, 3200000, 1, '台北辦公大樓A棟', '欣達冷凍', '含安裝調試'),
    ('變頻離心式冰水主機', '空調設備', '約克', 'YLAA0470', '1000RT 380V 60Hz', '台', 2022, 2950000, 1, '台北辦公大樓B棟', '欣達冷凍', ''),
    ('變頻離心式冰水主機', '空調設備', '開利', '30XW-V', '800RT 380V 60Hz', '台', 2024, 3500000, 1, '新竹科技廠辦', '開利空調', ''),
    ('冷卻水塔', '空調設備', '馬可尼', 'FRP-1000RT', '1000RT 逆流式', '台', 2023, 480000, 2, '台北辦公大樓A棟', '馬可尼', ''),
    ('冷卻水塔', '空調設備', '馬可尼', 'FRP-1000RT', '1000RT 逆流式', '台', 2022, 450000, 2, '台北辦公大樓B棟', '馬可尼', ''),
    ('冷卻水泵', '空調設備', '荏原', 'HSHE100-200', '100HP 380V', '台', 2023, 185000, 3, '台北辦公大樓A棟', '荏原幫浦', ''),
    ('空氣處理機', '空調設備', '約克', 'AHU-30000CMH', '30000CMH', '台', 2023, 320000, 4, '台北辦公大樓A棟', '欣達冷凍', ''),
    ('高壓氣體絕緣開關(GIS)', '電力設備', 'ABB', 'ZX2-17.5', '17.5kV 1250A', '面', 2023, 850000, 1, '台北辦公大樓A棟', '川湖電機', ''),
    ('高壓氣體絕緣開關(GIS)', '電力設備', 'ABB', 'ZX2-17.5', '17.5kV 1250A', '面', 2022, 780000, 1, '台北辦公大樓B棟', '川湖電機', ''),
    ('乾式變壓器', '電力設備', '正昇', 'TR-2000KVA', '2000KVA 22.8/0.48kV', '台', 2023, 680000, 2, '台北辦公大樓A棟', '正昇電機', ''),
    ('乾式變壓器', '電力設備', '正昇', 'TR-2000KVA', '2000KVA 22.8/0.48kV', '台', 2024, 720000, 2, '新竹科技廠辦', '正昇電機', ''),
    ('低壓配電盤 MCC', '電力設備', '士林', 'MCC-480V', '480V 2000A', '面', 2023, 420000, 3, '台北辦公大樓A棟', '士林電機', ''),
    ('緊急發電機', '電力設備', '康明斯', 'C1000D5', '1000kVA 380V 60Hz', '台', 2023, 2800000, 1, '台北辦公大樓A棟', '康明斯台灣', '含ATS'),
    ('緊急發電機', '電力設備', '康明斯', 'C1000D5', '1000kVA 380V 60Hz', '台', 2022, 2600000, 1, '台北辦公大樓B棟', '康明斯台灣', ''),
    ('UPS不斷電系統', '電力設備', 'EATON', '9395-300', '300kVA 雙轉換', '套', 2023, 1850000, 2, '台北辦公大樓A棟', '飛瑞股份', ''),
    ('消防幫浦組', '消防設備', '荏原', 'HSHD150', '150HP 電動主泵', '套', 2023, 680000, 1, '台北辦公大樓A棟', '荏原幫浦', '含柴油備用泵'),
    ('消防幫浦組', '消防設備', '荏原', 'HSHD150', '150HP 電動主泵', '套', 2022, 620000, 1, '台北辦公大樓B棟', '荏原幫浦', ''),
    ('自動撒水系統', '消防設備', '台灣印', 'ESFR-K25', 'ESFR快速反應', '式', 2023, 3200000, 1, '新竹科技廠辦', '台灣消防', ''),
    ('氣體滅火系統', '消防設備', '新昌', 'FM200-100kg', 'HFC-227ea 100kg', '套', 2023, 480000, 4, '台北辦公大樓A棟', '新昌消防', '機房用'),
    ('生活給水泵', '給排水設備', '荏原', 'EVMSU-5', '5HP 加壓泵', '台', 2023, 85000, 2, '台北辦公大樓A棟', '荏原幫浦', ''),
    ('污水處理設備', '給排水設備', '日揚', 'STP-200T', '200噸/日', '套', 2023, 2400000, 1, '新竹科技廠辦', '日揚環保', ''),
    ('門禁系統', '弱電系統', '台灣Suprema', 'BioEntry', '臉部辨識 門禁', '套', 2023, 1200000, 1, '台北辦公大樓A棟', '統寶弱電', '含軟體'),
    ('監控CCTV系統', '弱電系統', 'AXIS', 'P3245-V', '4MP 室內半球', '套', 2023, 850000, 1, '台北辦公大樓A棟', '統寶弱電', '含NVR及軟體'),
    ('電梯', '電梯/電扶梯', '三菱', 'NEXIEZ-MRL', '1350kg 2.5m/s', '台', 2023, 3800000, 6, '台北辦公大樓A棟', '三菱電梯', '無機房型'),
    ('電梯', '電梯/電扶梯', '三菱', 'NEXIEZ-MRL', '1350kg 2.5m/s', '台', 2022, 3500000, 4, '台北辦公大樓B棟', '三菱電梯', ''),
    ('電扶梯', '電梯/電扶梯', '奧的斯', 'XO-500A', '寬幅1000mm 0.5m/s', '台', 2023, 1800000, 4, '台北辦公大樓A棟', '奧的斯', ''),
]


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            brand TEXT DEFAULT '',
            model TEXT DEFAULT '',
            spec TEXT DEFAULT '',
            unit TEXT DEFAULT '式',
            year INTEGER NOT NULL,
            amount REAL NOT NULL,
            quantity REAL DEFAULT 1,
            project TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    count = conn.execute('SELECT COUNT(*) FROM items').fetchone()[0]
    if count == 0:
        conn.executemany('''
            INSERT INTO items (name, category, brand, model, spec, unit, year, amount, quantity, project, supplier, notes)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', SAMPLE_DATA)
    conn.commit()
    conn.close()


def build_filter_query(args):
    query = 'SELECT * FROM items WHERE 1=1'
    params = []

    if args.get('keyword', '').strip():
        kw = f"%{args['keyword'].strip()}%"
        query += ' AND (name LIKE ? OR spec LIKE ? OR model LIKE ? OR brand LIKE ? OR project LIKE ? OR notes LIKE ?)'
        params.extend([kw] * 6)

    if args.get('name', '').strip():
        query += ' AND name LIKE ?'
        params.append(f"%{args['name'].strip()}%")

    if args.get('category', '').strip():
        query += ' AND category = ?'
        params.append(args['category'].strip())

    if args.get('year_from', '').strip():
        query += ' AND year >= ?'
        params.append(int(args['year_from']))

    if args.get('year_to', '').strip():
        query += ' AND year <= ?'
        params.append(int(args['year_to']))

    if args.get('amount_min', '').strip():
        query += ' AND amount >= ?'
        params.append(float(args['amount_min']))

    if args.get('amount_max', '').strip():
        query += ' AND amount <= ?'
        params.append(float(args['amount_max']))

    if args.get('project', '').strip():
        query += ' AND project LIKE ?'
        params.append(f"%{args['project'].strip()}%")

    return query, params


@app.route('/')
def index():
    return render_template('index.html', categories=CATEGORIES)


@app.route('/api/items', methods=['GET'])
def get_items():
    query, params = build_filter_query(request.args)
    query += ' ORDER BY year DESC, name ASC'
    conn = get_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/items', methods=['POST'])
def create_item():
    data = request.json
    conn = get_db()
    conn.execute('''
        INSERT INTO items (name, category, brand, model, spec, unit, year, amount, quantity, project, supplier, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        data['name'], data['category'], data.get('brand', ''), data.get('model', ''),
        data.get('spec', ''), data.get('unit', '式'), int(data['year']), float(data['amount']),
        float(data.get('quantity', 1)), data.get('project', ''), data.get('supplier', ''), data.get('notes', '')
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/items/<int:item_id>', methods=['PUT'])
def update_item(item_id):
    data = request.json
    conn = get_db()
    conn.execute('''
        UPDATE items SET name=?, category=?, brand=?, model=?, spec=?, unit=?, year=?,
        amount=?, quantity=?, project=?, supplier=?, notes=? WHERE id=?
    ''', (
        data['name'], data['category'], data.get('brand', ''), data.get('model', ''),
        data.get('spec', ''), data.get('unit', '式'), int(data['year']), float(data['amount']),
        float(data.get('quantity', 1)), data.get('project', ''), data.get('supplier', ''), data.get('notes', ''),
        item_id
    ))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/items/<int:item_id>', methods=['DELETE'])
def delete_item(item_id):
    conn = get_db()
    conn.execute('DELETE FROM items WHERE id=?', (item_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/stats')
def get_stats():
    query, params = build_filter_query(request.args)
    base = query.replace('SELECT *', 'SELECT COUNT(*) as cnt, SUM(amount*quantity) as total, AVG(amount) as avg_amt', 1)
    conn = get_db()
    row = conn.execute(base, params).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route('/api/trend')
def get_trend():
    name = request.args.get('name', '').strip()
    category = request.args.get('category', '').strip()
    conn = get_db()
    query = 'SELECT year, AVG(amount) as avg_amount, MIN(amount) as min_amount, MAX(amount) as max_amount, COUNT(*) as count FROM items WHERE 1=1'
    params = []
    if name:
        query += ' AND name LIKE ?'
        params.append(f'%{name}%')
    if category:
        query += ' AND category = ?'
        params.append(category)
    query += ' GROUP BY year ORDER BY year'
    rows = conn.execute(query, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/projects')
def get_projects():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT project FROM items WHERE project != '' ORDER BY project").fetchall()
    conn.close()
    return jsonify([r['project'] for r in rows])


@app.route('/api/export')
def export_excel():
    query, params = build_filter_query(request.args)
    query += ' ORDER BY year DESC, category ASC, name ASC'
    conn = get_db()
    rows = conn.execute(query, params).fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '設備材料規格金額'

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='微軟正黑體', size=11)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['#', '年度', '類別', '名稱', '品牌', '型號', '規格', '單位', '金額(元)', '數量', '小計(元)', '專案', '供應商', '備註']
    col_widths = [5, 8, 12, 22, 12, 16, 30, 6, 14, 8, 14, 20, 14, 20]
    keys = ['id', 'year', 'category', 'name', 'brand', 'model', 'spec', 'unit', 'amount', 'quantity', None, 'project', 'supplier', 'notes']

    ws.row_dimensions[1].height = 28
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    alt_fill = PatternFill(start_color='E8F0FE', end_color='E8F0FE', fill_type='solid')
    num_fmt = '#,##0'

    for row_idx, row in enumerate(rows, 2):
        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, key in enumerate(keys, 1):
            if key is None:
                val = (row['amount'] or 0) * (row['quantity'] or 1)
            else:
                val = row[key]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col_idx == 7))
            if fill:
                cell.fill = fill
            if col_idx in (9, 10, 11):
                cell.number_format = num_fmt

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'ME設備材料_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/template')
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '匯入範本'

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, name='微軟正黑體')
    note_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    for col, h in enumerate(IMPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    samples = [
        ['變頻離心式冰水主機', '空調設備', 2024, 3200000, '約克', 'YLAA0470', '1000RT 380V 60Hz', '台', 1, '台北辦公大樓', '欣達冷凍', '含安裝'],
        ['乾式變壓器', '電力設備', 2024, 720000, '正昇', 'TR-2000KVA', '2000KVA 22.8/0.48kV', '台', 2, '新竹廠辦', '正昇電機', ''],
        ['消防幫浦組', '消防設備', 2024, 680000, '荏原', 'HSHD150', '150HP 電動主泵', '套', 1, '台北辦公大樓', '荏原幫浦', '含柴油備用泵'],
    ]
    for r_idx, row in enumerate(samples, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    note_row = len(samples) + 3
    note_cell = ws.cell(row=note_row, column=1, value='說明：* 為必填欄位。類別請填入：空調設備、電力設備、給排水設備、消防設備、弱電系統、機械設備、電梯/電扶梯、其他')
    note_cell.fill = note_fill
    note_cell.font = Font(color='7F6000')
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=12)

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name='ME_匯入範本.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/import', methods=['POST'])
def import_excel():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': '未收到檔案'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'無法開啟 Excel 檔案：{e}'}), 400

    conn = get_db()
    count = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):
            continue
        try:
            name = str(row[0] or '').strip()
            category = str(row[1] or '').strip()
            year = int(row[2]) if row[2] else datetime.now().year
            amount = float(row[3]) if row[3] is not None else 0
            brand = str(row[4] or '').strip()
            model_val = str(row[5] or '').strip()
            spec = str(row[6] or '').strip()
            unit = str(row[7] or '式').strip()
            quantity = float(row[8]) if row[8] is not None else 1
            project = str(row[9] or '').strip()
            supplier = str(row[10] or '').strip()
            notes = str(row[11] or '').strip()

            if not name:
                errors.append(f'第 {row_idx} 行：名稱為空，略過')
                continue
            if not category:
                errors.append(f'第 {row_idx} 行：類別為空，略過')
                continue

            conn.execute('''
                INSERT INTO items (name, category, brand, model, spec, unit, year, amount, quantity, project, supplier, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (name, category, brand, model_val, spec, unit, year, amount, quantity, project, supplier, notes))
            count += 1
        except Exception as e:
            errors.append(f'第 {row_idx} 行：{e}')

    conn.commit()
    conn.close()
    return jsonify({'imported': count, 'errors': errors})


if __name__ == '__main__':
    init_db()
    print('機電工程設備材料規格金額查詢系統')
    print('請開啟瀏覽器前往 http://127.0.0.1:5000')
    app.run(debug=True, port=5000)
